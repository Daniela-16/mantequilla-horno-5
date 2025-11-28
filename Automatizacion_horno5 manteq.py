# -*- coding: utf-8 -*-
"""
Creado el Lunes 24 de Noviembre de 2025

@author: NCGNpracpim
"""

import pandas as pd
import numpy as np
import streamlit as st
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import Counter
import re
from typing import Tuple, Union, Dict, Any

# --- NOMBRES DE COLUMNAS CLAVE Y CONSTANTES COMUNES ---
COL_CANT_CALCULADA = 'Cant. base calculada'
COL_PESO_NETO = 'peso neto'
COL_SECUENCIA = 'secuencia recurso'
COL_ATIPICO = 'Atipico_Cant_Calculada'
COL_MANO_OBRA = 'Mano de obra'
HOJA_MANO_OBRA = 'Mano de obra' # Esta hoja es común
COL_SUMA_VALORES = 'suma valores'
COL_PORCENTAJE_RECHAZO = '%de rechazo'
COL_NRO_PERSONAS = 'Cant_Manual'
COL_NRO_MAQUINAS = 'Cant_Maquinas'
COL_CLAVE = 'Clave_Busqueda'
COL_DIFERENCIA = 'diferencia'
NOMBRE_COL_CANTIDAD_BASE = 'Cantidad base'
NOMBRE_COL_CLAVE_EXTERNA = 'MaterialHorno'
NOMBRE_COL_CANT_EXTERNA = 'CantidadBaseXHora'
# NUEVA CONSTANTE PARA EL CAMPO LINEA (Columna T)
COL_LINEA = 'Linea'
COL_PSTTBJO_CONCATENADO = 'PstoTbjo_Concat'

# Nombres de hojas a crear (Comunes)
HOJA_SECUENCIAS = 'Secuencias' # Esta hoja es común
HOJA_LSMW = 'lsmw'
HOJA_CAMPOS_USUARIO = 'campos de usuario'
HOJA_PORCENTAJE_RECHAZO = '% de rechazo'

# Columnas a resaltar en todas las hojas (solicitado por el usuario)
COLUMNAS_A_RESALTAR = [
    COL_MANO_OBRA,
    COL_SUMA_VALORES,
    COL_NRO_PERSONAS,
    COL_NRO_MAQUINAS
]

# Definición de columnas de salida (Comunes)
COLUMNAS_LSMW = [
    'PstoTbjo', 'GrpHRuta', 'CGH', 'Material', 'Ce.', 'Op.',
    COL_CANT_CALCULADA, 'ValPref', 'ValPref1', COL_MANO_OBRA, 'ValPref3',
    COL_SUMA_VALORES, 'ValPref5'
]
COLUMNAS_CAMPOS_USUARIO = [
    'GrpHRuta', 'CGH', 'Material', 'Ce.', 'Op.',
    'Indicador', 'clase de control',
    COL_NRO_PERSONAS, COL_NRO_MAQUINAS
]
COLUMNAS_RECHAZO = [
    'GrPlf', 'Clave_Busqueda', 'Material', 'Ce.', 'alternativa', 'alternativa',
    'posición', 'Relevancia', COL_PORCENTAJE_RECHAZO,
    '% rechazo anterior', 'Diferencia', 'Txt.brv.HRuta'
]

# --- CONSTANTES ESPECÍFICAS DE CADA HORNO ---
HORNOS_CONFIG = {
    'HORNO 5': {
        'HOJA_PRINCIPAL': 'HORNO 5',
        'HOJA_SALIDA': 'HORNO5_procesado',
    },
    'HORNO 12': {
        'HOJA_PRINCIPAL': 'HORNO 12',
        'HOJA_SALIDA': 'HORNO12_procesado',
    },
    'HORNO 1': {
        'HOJA_PRINCIPAL': 'HORNO 1',
        'HOJA_SALIDA': 'HORNO1_procesado',
    },
    'HORNO 3': {
        'HOJA_PRINCIPAL': 'HORNO 3',
        'HOJA_SALIDA': 'HORNO3_procesado',
    },
}

# Índices para el archivo original (ASUMO QUE SON COMUNES)
IDX_MATERIAL = 2 # Columna C
IDX_GRPLF = 4 # Columna E
IDX_PSTTBJO = 18 # Columna S (Puesto de Trabajo)
IDX_LINEA = 19 # Columna T (Nueva)
IDX_CANTIDAD_BASE = 6
IDX_MATERIAL_PN = 0
IDX_RECHAZO_EXTERNA = 28

# --- FUNCIONES DE LÓGICA ---

def detectar_y_marcar_cantidad_atipica(grupo: pd.DataFrame) -> pd.Series:
    """Identifica valores atípicos (diferentes de la moda) en Cant. base calculada dentro de un grupo."""
    valores_no_nan = grupo[COL_CANT_CALCULADA].dropna()
    if valores_no_nan.empty:
        return pd.Series(False, index=grupo.index)

    conteo = Counter(valores_no_nan)
    moda = conteo.most_common(1)[0][0]

    es_diferente_a_moda = grupo[COL_CANT_CALCULADA] != moda

    return es_diferente_a_moda

def crear_y_guardar_hoja(wb, df_base: pd.DataFrame, nombre_hoja: str, columnas_destino: list, fill_encabezado: PatternFill, font_negrita: Font):
    """
    Crea una nueva hoja, la rellena con las columnas especificadas de df_base,
    y aplica formato de encabezado a las columnas definidas en COLUMNAS_A_RESALTAR.
    """
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]

    ws = wb.create_sheet(nombre_hoja)

    # 1. Crear el nuevo DataFrame con las columnas solicitadas
    df_nuevo = pd.DataFrame()
    for col in columnas_destino:
        df_nuevo[col] = df_base[col] if col in df_base.columns else np.nan

    # 2. Escribir el nuevo DataFrame en la hoja
    for row in dataframe_to_rows(df_nuevo, header=True, index=False):
        ws.append(row)

    # 3. Aplicar Formato a Encabezados Específicos
    indices_a_formatear = [
        df_nuevo.columns.get_loc(col) + 1 
        for col in COLUMNAS_A_RESALTAR 
        if col in df_nuevo.columns
    ]

    for col_idx in indices_a_formatear:
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.fill = fill_encabezado
        header_cell.font = font_negrita

def obtener_secuencia(puesto_trabajo: str, df_secuencias: pd.DataFrame) -> Union[int, float]:
    """Busca la secuencia del puesto de trabajo en la hoja 'Secuencias'."""
    psttbjo_str = str(puesto_trabajo).strip()

    for col_idx in range(df_secuencias.shape[1]):
        col_data = df_secuencias.iloc[:, col_idx].dropna().astype(str).str.strip()
        psttbjo_sec = set(col_data)

        if psttbjo_str in psttbjo_sec:
            return col_idx + 1

    return np.nan

def cargar_y_limpiar_datos(file_original: io.BytesIO, file_info_externa: io.BytesIO, nombre_horno: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, str]]:
    """Carga todos los DataFrames necesarios desde los buffers de archivo."""
    
    config = HORNOS_CONFIG[nombre_horno]
    hoja_principal = config['HOJA_PRINCIPAL']

    # --- 1. Lectura de Archivo Original ---
    # Leer encabezados de la hoja principal (para obtener nombres de columnas por índice)
    cols_original = pd.read_excel(file_original, sheet_name=hoja_principal, nrows=0).columns.tolist()
    file_original.seek(0)
    
    # Leer encabezados de Peso neto (para obtener nombres de columnas por índice)
    cols_pn = pd.read_excel(file_original, sheet_name='Peso neto', nrows=0).columns.tolist()
    file_original.seek(0)
    
    col_names = {
        'cant_base_leida': cols_original[IDX_CANTIDAD_BASE],
        'material': cols_original[IDX_MATERIAL],
        'psttbjo': cols_original[IDX_PSTTBJO],
        'material_pn': cols_pn[IDX_MATERIAL_PN],
        'peso_neto_valor': cols_pn[2],
    }

    # Definir las columnas a leer de la hoja principal
    usecols_original = list(range(len(cols_original)))
    
    # Si es HORNO 1, aseguramos que se lea la columna T (Línea)
    if nombre_horno == 'HORNO 1':
        if IDX_LINEA >= len(cols_original):
            # Asumimos que la columna existe y extendemos usecols
            usecols_original.append(IDX_LINEA)
        # Añadir el nombre de la columna Línea a col_names si es necesario
        if IDX_LINEA < len(cols_original):
            col_names[COL_LINEA] = cols_original[IDX_LINEA]
        else:
            # Si el archivo no tiene la columna T, asignamos un nombre temporal
            col_names[COL_LINEA] = 'Columna T (Linea)'
    
    # Carga de DataFrames
    df_original = pd.read_excel(
        file_original, 
        sheet_name=hoja_principal, 
        dtype={col_names['cant_base_leida']: str},
        usecols=usecols_original # Leer las columnas definidas
    )
    file_original.seek(0)
    
    # Si la columna 'Linea' no se leyó correctamente o no existía, la creamos vacía para evitar errores
    if nombre_horno == 'HORNO 1' and COL_LINEA not in df_original.columns:
        # Si la columna 19 no se leyó o no tenía encabezado, intentamos forzar el nombre
        # Esto es un parche si el encabezado de la columna T es nulo
        if IDX_LINEA < len(cols_original) and cols_original[IDX_LINEA] not in df_original.columns:
            df_original.rename(columns={df_original.columns[IDX_LINEA]: COL_LINEA}, inplace=True)
        elif COL_LINEA not in df_original.columns:
             # Si no se pudo leer con el nombre, la inicializamos a NaN
            df_original[COL_LINEA] = np.nan
            
        # Re-actualizar cols_original si se renombró o si se añadió la columna
        cols_original = df_original.columns.tolist()
        if COL_LINEA in df_original.columns:
             col_names[COL_LINEA] = COL_LINEA


    df_peso_neto = pd.read_excel(file_original, sheet_name='Peso neto')
    file_original.seek(0)

    df_secuencias = pd.read_excel(file_original, sheet_name=HOJA_SECUENCIAS)
    file_original.seek(0)
    
    # ******* CORRECCIÓN PARA LA HOJA MANO DE OBRA *******
    # Forzamos la lectura de 5 columnas (0 a 4) para evitar el error 'list index out of range' 
    # si la Columna E (Cant_Personas) no tiene datos en las primeras filas.
    columnas_mano_obra = [0, 1, 2, 3, 4] # Índices esperados
    
    df_mano_obra = pd.read_excel(
        file_original, 
        sheet_name=HOJA_MANO_OBRA, 
        header=None, 
        usecols=range(len(columnas_mano_obra)), # Lectura forzada de 5 columnas
        names=columnas_mano_obra # Nombramos las columnas con los índices esperados
    )
    file_original.seek(0)
    # ******* FIN DE CORRECCIÓN *******

    # --- 2. Lectura de Archivo Externo ---
    cols_externo = pd.read_excel(file_info_externa, sheet_name='Especif y Rutas', nrows=0).columns.tolist()
    file_info_externa.seek(0)

    nombre_col_rechazo_externa = cols_externo[IDX_RECHAZO_EXTERNA] if IDX_RECHAZO_EXTERNA < len(cols_externo) else 'Columna AC'
    cols_a_leer_externo = [NOMBRE_COL_CLAVE_EXTERNA, NOMBRE_COL_CANT_EXTERNA, nombre_col_rechazo_externa]
    df_externo = pd.read_excel(file_info_externa, sheet_name='Especif y Rutas', header=0, usecols=cols_a_leer_externo)
    file_info_externa.seek(0)

    # Renombrar columna de cantidad base si es necesario
    if col_names['cant_base_leida'] != NOMBRE_COL_CANTIDAD_BASE:
        df_original = df_original.rename(columns={col_names['cant_base_leida']: NOMBRE_COL_CANTIDAD_BASE})

    # Guardar nombres de columnas externas leídas
    col_names['nombre_col_rechazo_externa'] = nombre_col_rechazo_externa
    col_names['cols_original'] = cols_original
    col_names['hoja_principal'] = hoja_principal

    return df_original, df_externo, df_peso_neto, df_secuencias, df_mano_obra, col_names

# --- FUNCIÓN PRINCIPAL DE PROCESAMIENTO ---

def automatizacion_final_diferencia_reforzada(file_original: io.BytesIO, file_info_externa: io.BytesIO, nombre_horno: str) -> Tuple[bool, Union[str, io.BytesIO]]:
    """
    Ejecuta toda la lógica de procesamiento.
    """
    
    config = HORNOS_CONFIG[nombre_horno]
    HOJA_SALIDA = config['HOJA_SALIDA']
    
    FINAL_COL_ORDER = [
        'GrpHRuta', 'CGH', 'Material', COL_CLAVE, 'Ce.', 'GrPlf', 'Op.',
        COL_PORCENTAJE_RECHAZO, NOMBRE_COL_CANTIDAD_BASE, COL_CANT_CALCULADA,
        COL_DIFERENCIA, COL_PESO_NETO, COL_SECUENCIA, 'ValPref', 'ValPref1',
        'ValPref2', COL_MANO_OBRA, 'ValPref3', 'ValPref4', COL_SUMA_VALORES,
        'ValPref5',
        'Campo de usuario cantidad MANUAL',
        COL_NRO_PERSONAS,
        'Campo de usuario cantidad MAQUINAS',
        COL_NRO_MAQUINAS,
        'Texto breve operación', 'Ctrl', 'VerF', 'PstoTbjo', 'Cl.', 'Gr.fam.pre',
        'Texto breve de material', 'Txt.brv.HRuta', 'Bloq.vers.fabric.', 'Campo usuario unidad',
        'Campo usuario unidad.1', 'Cantidad', 'Contador', 'InBo', 'InBo.1', 'InBo.2',
        'Unnamed: 31', 'I'
    ]
    
    COLUMNAS_A_SUMAR = ['ValPref', 'ValPref1', COL_MANO_OBRA, 'ValPref3']

    try:
        st.write("---")
        st.subheader(f"Preparando datos para **{nombre_horno}**... 📊")

        # 1. Carga y limpieza de datos (Se pasa el nombre del horno)
        df_original, df_externo, df_peso_neto, df_secuencias, df_mano_obra, col_names = cargar_y_limpiar_datos(file_original, file_info_externa, nombre_horno)

        # 2. Creación de la Clave de Búsqueda
        def limpiar_col(df: pd.DataFrame, col_name: str) -> pd.Series:
            """Limpia (quita caracteres no alfanuméricos) la columna especificada."""
            if col_name not in df.columns:
                raise KeyError(f"Columna '{col_name}' no encontrada en la hoja '{col_names['hoja_principal']}'.")
            return df[col_name].astype(str).str.strip().str.replace(r'\W+', '', regex=True)

        # Usamos los nombres de columna reales de df_original, no los índices directos
        material_col_name = col_names['material']
        grplf_col_name = col_names['cols_original'][IDX_GRPLF]
        psttbjo_col_name = col_names['psttbjo']

        df_original[COL_CLAVE] = (
            limpiar_col(df_original, material_col_name) +
            limpiar_col(df_original, grplf_col_name) +
            limpiar_col(df_original, psttbjo_col_name)
        )

        df_externo[NOMBRE_COL_CLAVE_EXTERNA] = df_externo[NOMBRE_COL_CLAVE_EXTERNA].astype(str).str.strip().str.replace(r'\W+', '', regex=True)

        # --- LÓGICA ESPECÍFICA DE HORNO 1 PARA SECUENCIA ---
        columna_para_secuencia = psttbjo_col_name
        
        if nombre_horno == 'HORNO 1':
            st.info("⚠️ Aplicando lógica especial: La secuencia del HORNO 1 se calcula con PstoTbjo y Línea.")
            
            # Limpiar la columna de línea y reemplazar NaN con cadena vacía para la concatenación
            linea_limpia = df_original.get(COL_LINEA, pd.Series([''] * len(df_original))).astype(str).str.strip().replace('', np.nan)
            psttbjo_limpio = df_original[psttbjo_col_name].astype(str).str.strip()
            
            # Crear la columna concatenada (PstoTbjo si Línea es NaN/vacío, PstoTbjo + Línea si no)
            df_original[COL_PSTTBJO_CONCATENADO] = np.where(
                pd.isna(linea_limpia),
                psttbjo_limpio,
                psttbjo_limpio + linea_limpia
            )
            columna_para_secuencia = COL_PSTTBJO_CONCATENADO
        # ---------------------------------------------------

        # 3. Mapeo de Cantidad Calculada, Rechazo y Peso Neto
        def mapear_columna(df_mapeo: pd.DataFrame, col_indice: str, col_destino: str, col_clave: str, nombre_col_mapa: str):
            """Realiza el mapeo de una columna externa al DataFrame principal."""
            mapa = df_mapeo.drop_duplicates(subset=[col_clave], keep='first').set_index(col_clave)[nombre_col_mapa]
            df_original[col_destino] = df_original[col_indice].map(mapa)

        mapear_columna(df_externo, COL_CLAVE, COL_CANT_CALCULADA, NOMBRE_COL_CLAVE_EXTERNA, NOMBRE_COL_CANT_EXTERNA)
        mapear_columna(df_externo, COL_CLAVE, COL_PORCENTAJE_RECHAZO, NOMBRE_COL_CLAVE_EXTERNA, col_names['nombre_col_rechazo_externa'])
        mapear_columna(df_peso_neto, material_col_name, COL_PESO_NETO, col_names['material_pn'], col_names['peso_neto_valor'])

        # 4. Cálculo de Secuencia
        # Usamos la columna condicionalmente seleccionada
        df_original[COL_SECUENCIA] = df_original[columna_para_secuencia].apply(lambda x: obtener_secuencia(x, df_secuencias))

        # 5. Cálculo de Mano de Obra, Personas y Máquinas
        # Índices de df_mano_obra (A=0, C=2, D=3, E=4)
        COL_PSTTBJO_MO = 0 
        COL_TIEMPO_MO = 2 
        COL_CANTIDAD_MAQUINAS_MO = 3 
        COL_CANTIDAD_PERSONAS_MO = 4 # Ahora sabemos que este índice existe gracias a la corrección

        # Limpieza de datos en df_mano_obra
        df_mano_obra[COL_PSTTBJO_MO] = df_mano_obra[COL_PSTTBJO_MO].astype(str).str.strip()
        for col_idx in [COL_TIEMPO_MO, COL_CANTIDAD_MAQUINAS_MO, COL_CANTIDAD_PERSONAS_MO]:
            # Convertimos a numérico, los NaNs generados por la lectura de celdas vacías se manejan aquí.
            df_mano_obra[col_idx] = pd.to_numeric(df_mano_obra[col_idx], errors='coerce') 

        # Filtro: Solo operaciones que terminan en '1'
        COL_OP = 'Op.'
        op_col = df_original[COL_OP].astype(str).str.strip()
        indices_terminan_en_1 = op_col.str.endswith('1')
        psttbjo_filtrado = df_original.loc[indices_terminan_en_1, psttbjo_col_name].astype(str).str.strip() # Usamos el PstoTbjo original

        # Mapeos para Mano de Obra, Personas y Máquinas
        def mapear_mo_filtros(col_origen: int, col_destino: str):
            """Genera el mapa y aplica el mapeo solo a las filas filtradas."""
            mapa = df_mano_obra.drop_duplicates(subset=[COL_PSTTBJO_MO], keep='first').set_index(COL_PSTTBJO_MO)[col_origen]
            df_original.loc[indices_terminan_en_1, col_destino] = psttbjo_filtrado.map(mapa)

        # 5.1. Tiempo de Mano de Obra (Personas * 60)
        df_mano_obra['Calculo_MO_Tiempo'] = df_mano_obra[COL_TIEMPO_MO] * 60
        mapa_mano_obra_tiempo = df_mano_obra.drop_duplicates(subset=[COL_PSTTBJO_MO], keep='first').set_index(COL_PSTTBJO_MO)['Calculo_MO_Tiempo']
        df_original[COL_MANO_OBRA] = np.nan
        df_original.loc[indices_terminan_en_1, COL_MANO_OBRA] = psttbjo_filtrado.map(mapa_mano_obra_tiempo)

        # 5.2. Número de Personas (Columna E)
        df_original[COL_NRO_PERSONAS] = np.nan
        mapear_mo_filtros(COL_CANTIDAD_PERSONAS_MO, COL_NRO_PERSONAS)

        # 5.3. Número de Máquinas (Columna D)
        df_original[COL_NRO_MAQUINAS] = np.nan
        mapear_mo_filtros(COL_CANTIDAD_MAQUINAS_MO, COL_NRO_MAQUINAS)

        # 6. Suma de Valores y formato
        def formato_excel_regional_suma(x):
            """Aplica formato de coma decimal para Excel y maneja NaN/cero."""
            return f"{x:.2f}".replace('.', ',') if pd.notna(x) and x != 0.0 else np.nan

        df_temp_sum = df_original[COLUMNAS_A_SUMAR].apply(lambda col: pd.to_numeric(col, errors='coerce'))
        df_original[COL_SUMA_VALORES] = df_temp_sum.sum(axis=1, skipna=True).apply(formato_excel_regional_suma)

        # 7. Cálculo de Diferencia y Atípicos
        H = pd.to_numeric(df_original[NOMBRE_COL_CANTIDAD_BASE].astype(str).str.replace(',', '.', regex=False).str.strip(), errors='coerce')
        I = pd.to_numeric(df_original[COL_CANT_CALCULADA], errors='coerce')

        diferencia_calculada = H.fillna(0) - I.fillna(0)

        def formato_excel_regional(x):
            """Aplica formato de coma decimal para Excel."""
            return f"{x:.2f}".replace('.', ',') if pd.notna(x) else np.nan

        df_original[COL_DIFERENCIA] = diferencia_calculada.apply(formato_excel_regional)

        # Atípicos
        df_original[COL_CANT_CALCULADA] = I 
        cols_agrupamiento = [COL_PESO_NETO, COL_SECUENCIA]
        for col in cols_agrupamiento:
            if col not in df_original.columns:
                 # Si falta una columna clave (esto solo debería ocurrir si los mapeos fallan)
                 raise KeyError(f"Columna de agrupamiento '{col}' falta en el DataFrame. Se necesita para calcular atípicos.")

        df_original[COL_ATIPICO] = df_original.groupby(cols_agrupamiento, dropna=True).apply(
            detectar_y_marcar_cantidad_atipica
        ).reset_index(level=[0, 1], drop=True).fillna(False)


        # 8. Reconstrucción Final y Guardado con Formato
        df_original_final = df_original.reindex(columns=[c for c in FINAL_COL_ORDER if c in df_original.columns])

        # Cargar el libro de trabajo desde el buffer (Para mantener las hojas originales)
        file_original.seek(0)
        wb = load_workbook(file_original)

        # Definición de Estilos
        fill_anomalia = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid') # Naranja (Atípicos)
        fill_encabezado = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid') # Azul claro (Calculadas)
        font_negrita = Font(bold=True)

        # Crear y escribir la hoja principal procesada
        if HOJA_SALIDA in wb.sheetnames: del wb[HOJA_SALIDA]
        ws = wb.create_sheet(HOJA_SALIDA)

        for row in dataframe_to_rows(df_original_final, header=True, index=False):
            ws.append(row)

        # 4. APLICACIÓN DE FORMATOS EN HOJA PRINCIPAL
        COLUMNAS_ENCABEZADO_FORMATO = [COL_CANT_CALCULADA] + COLUMNAS_A_RESALTAR

        indices_encabezado = [
            df_original_final.columns.get_loc(col_name) + 1 
            for col_name in COLUMNAS_ENCABEZADO_FORMATO
            if col_name in df_original_final.columns
        ]

        for col_idx in indices_encabezado:
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.fill = fill_encabezado
            header_cell.font = font_negrita

        # Aplicar el sombreado a los datos de 'Cant. base calculada' (naranja para atípicos)
        try:
            col_cant_calculada_idx = df_original_final.columns.get_loc(COL_CANT_CALCULADA) + 1
        except KeyError:
            col_cant_calculada_idx = 9 

        for r in range(2, len(df_original) + 2):
            if df_original.iloc[r-2][COL_ATIPICO]:
                cell_to_color = ws.cell(row=r, column=col_cant_calculada_idx)
                cell_to_color.fill = fill_anomalia

        # --- CREACIÓN DE HOJAS ADICIONALES (Se pasan los estilos) ---
        crear_y_guardar_hoja(wb, df_original, HOJA_LSMW, COLUMNAS_LSMW, fill_encabezado, font_negrita)
        crear_y_guardar_hoja(wb, df_original, HOJA_CAMPOS_USUARIO, COLUMNAS_CAMPOS_USUARIO, fill_encabezado, font_negrita)
        crear_y_guardar_hoja(wb, df_original, HOJA_PORCENTAJE_RECHAZO, COLUMNAS_RECHAZO, fill_encabezado, font_negrita)

        # Guardar el libro de trabajo modificado en un buffer de Bytes
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        return True, output_buffer

    except KeyError as ke:
        return False, f"❌ ERROR CRÍTICO DE ENCABEZADO: El script no encontró la columna {ke}. Verifique las hojas y encabezados del archivo original o externo. Asegúrese que el nombre de la hoja principal **{config['HOJA_PRINCIPAL']}** es correcto."
    except IndexError as ie:
        # Este error es menos probable con la corrección, pero lo mantenemos por si acaso
        return False, f"❌ ERROR CRÍTICO DE ÍNDICE: Un índice de columna está fuera de rango. Mensaje: {ie}"
    except ValueError as ve:
        if 'sheetname' in str(ve) or 'Worksheet' in str(ve):
            hojas_requeridas = [config['HOJA_PRINCIPAL'], 'Peso neto', HOJA_SECUENCIAS, HOJA_MANO_OBRA, 'Especif y Rutas']
            return False, f"❌ Error de Lectura de Hoja: Una de las hojas clave ({', '.join(hojas_requeridas)}) no se encontró en los archivos cargados. Mensaje: {ve}"
        return False, f"❌ Ocurrió un error inesperado de valor. Mensaje: {ve}"
    except Exception as e:
        return False, f"❌ Ocurrió un error inesperado. Mensaje: {e}"


# --- INTERFAZ DE STREAMLIT (CON SELECTOR DE HORNO) ---

def main():
    """Configura la interfaz de usuario de Streamlit."""
    st.set_page_config(
        page_title="Automatización Hornos",
        layout="centered",
        initial_sidebar_state="auto"
    )

    st.title("⚙️ Automatización Verificación de datos - HORNOS")
    st.markdown("Seleccione el Horno a procesar y luego cargue los archivos.")

    # SELECCIÓN DEL HORNO
    hornos_disponibles = list(HORNOS_CONFIG.keys())
    selected_horno = st.radio(
        "**1. Seleccione el Horno a Procesar:**",
        hornos_disponibles,
        index=hornos_disponibles.index('HORNO 5') if 'HORNO 5' in hornos_disponibles else 0,
        horizontal=True
    )
    st.markdown("---")
    
    config = HORNOS_CONFIG[selected_horno]
    hoja_principal = config['HOJA_PRINCIPAL']
    hoja_salida = config['HOJA_SALIDA']

    st.subheader(f"2. Carga de Archivos para **{selected_horno}** (Hoja Principal: '{hoja_principal}')")
    
    col1, col2 = st.columns(2)

    with col1:
        file_original = st.file_uploader(
            f"Carga la base de datos original",
            type=['xlsx'],
            help=f"El archivo debe contener las hojas: **{hoja_principal}**, 'Peso neto', '{HOJA_SECUENCIAS}' y '{HOJA_MANO_OBRA}'."
        )

    with col2:
        file_externa = st.file_uploader(
            "Carga el archivo externo de toma de información.",
            type=['xlsb', 'xlsx'],
            help="El archivo que contiene la hoja 'Especif y Rutas'."
        )

    st.markdown("---")

    # Botón de ejecución y manejo del proceso
    if st.button(f"▶️ PROCESAR {selected_horno}", type="primary", use_container_width=True):
        if file_original is None or file_externa is None:
            st.error("Por favor, cargue ambos archivos antes de procesar.")
        else:
            file_buffer_original = io.BytesIO(file_original.getvalue())
            file_buffer_externa = io.BytesIO(file_externa.getvalue())

            with st.spinner(f'Procesando datos y generando reporte para {selected_horno}...'):
                success, resultado = automatizacion_final_diferencia_reforzada(
                    file_buffer_original,
                    file_buffer_externa,
                    selected_horno
                )

            st.markdown("---")

            if success:
                st.success(f"✅ Proceso para **{selected_horno}** completado exitosamente.")

                # Botón de Descarga
                # Intentamos crear un nombre de archivo limpio
                base_name = file_original.name.split('.')[0]
                if hoja_principal in base_name:
                    file_name_output = base_name.replace(hoja_principal, '') + f"{hoja_salida}.xlsx"
                else:
                    file_name_output = f"{base_name}_{hoja_salida}.xlsx"
                
                st.download_button(
                    label="⬇️ Descargar Archivo Procesado",
                    data=resultado,
                    file_name=file_name_output,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True
                )
                st.info(f"El archivo descargado contiene todas las hojas originales más las 4 hojas de reporte: **{hoja_salida}**, '{HOJA_LSMW}', '{HOJA_CAMPOS_USUARIO}' y '{HOJA_PORCENTAJE_RECHAZO}'.")
            else:
                st.error("❌ Error en el Proceso")
                st.warning(resultado)
                st.write("Verifique el formato de las hojas y los nombres de las columnas en sus archivos.")

if __name__ == "__main__":
    main()
if __name__ == "__main__":
    main()

