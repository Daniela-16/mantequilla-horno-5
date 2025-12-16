import pandas as pd
import numpy as np
import streamlit as st
import io
import re
from typing import Tuple, Union, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from collections import Counter

# --- 1. CONSTANTES CENTRALIZADAS Y DEFINICIONES ---
COL = {
    'CANT_CALCULADA': 'Cant. base calculada',
    'PESO_NETO': 'peso neto',
    'SECUENCIA': 'secuencia recurso',
    'ATIPICO': 'Atipico_Cant_Calculada',
    'MANO_OBRA': 'Mano de obra',
    'SUMA_VALORES': 'suma valores',
    'PORCENTAJE_RECHAZO': '%de rechazo',
    'NRO_PERSONAS': 'Cant_Manual',
    'NRO_MAQUINAS': 'Cant_Maquinas',
    'CLAVE_BUSQUEDA': 'Clave_Busqueda',
    'DIFERENCIA': 'diferencia',
    'CANTIDAD_BASE': 'Cantidad base',
    'OP': 'Op.',
    'LINEA': 'Linea',
    'CLAVE_EXTERNA': 'MaterialHorno',
    'CANT_EXTERNA': 'CantidadBaseXHora',
    'HOJA_SALIDA_SECUENCIAS': 'Secuencias',
    'HOJA_SALIDA_LSMW': 'lsmw',
    'HOJA_SALIDA_CAMPOS_USUARIO': 'campos de usuario',
    'HOJA_SALIDA_RECHAZO': '% de rechazo',
    'HOJA_MANO_OBRA': 'Mano de obra',
    'RESALTAR': ['Mano de obra', 'suma valores', 'Cant_Manual', 'Cant_Maquinas']
}

HORNOS_CONFIG = {
    'HORNO 1': {'HOJA_PRINCIPAL': 'HORNO 1', 'HOJA_SALIDA': 'HORNO1_procesado'},
    'HORNO 2': {'HOJA_PRINCIPAL': 'HORNO 2', 'HOJA_SALIDA': 'HORNO2_procesado'},
    'HORNO 3': {'HOJA_PRINCIPAL': 'HORNO 3', 'HOJA_SALIDA': 'HORNO3_procesado'},
    'HORNO 4': {'HOJA_PRINCIPAL': 'HORNO 4', 'HOJA_SALIDA': 'HORNO4_procesado'},
    'HORNO 5': {'HOJA_PRINCIPAL': 'HORNO 5', 'HOJA_SALIDA': 'HORNO5_procesado'},
    'HORNO 6': {'HOJA_PRINCIPAL': 'HORNO 6', 'HOJA_SALIDA': 'HORNO6_procesado'},
    'HORNO 7': {'HOJA_PRINCIPAL': 'HORNO 7', 'HOJA_SALIDA': 'HORNO7_procesado'},
    'HORNO 8': {'HOJA_PRINCIPAL': 'HORNO 8', 'HOJA_SALIDA': 'HORNO8_procesado'},
    'HORNO 9': {'HOJA_PRINCIPAL': 'HORNO 9', 'HOJA_SALIDA': 'HORNO9_procesado'},
    'HORNO 10': {'HOJA_PRINCIPAL': 'HORNO 10', 'HOJA_SALIDA': 'HORNO10_procesado'},
    'HORNO 11': {'HOJA_PRINCIPAL': 'HORNO 11', 'HOJA_SALIDA': 'HORNO11_procesado'},
    'HORNO 12': {'HOJA_PRINCIPAL': 'HORNO 12', 'HOJA_SALIDA': 'HORNO12_procesado'},
    'HORNO 18': {'HOJA_PRINCIPAL': 'HORNO 18', 'HOJA_SALIDA': 'HORNO18_procesado'},
}

IDX = {
    'MATERIAL': 2, 'GRPLF': 4, 'CANTIDAD_BASE_LEIDA': 6, 'PSTTBJO': 18,
    'MATERIAL_PN': 0, 'RECHAZO_EXTERNA': 28, 'PESO_NETO_VALOR': 2,
}

COLUMNAS_OUTPUT = {
    'LSMW': ['PstoTbjo', 'GrpHRuta', 'CGH', 'Material', COL['CLAVE_BUSQUEDA'], 'Ce.', COL['OP'],
             COL['CANT_CALCULADA'], 'ValPref', 'ValPref1', COL['MANO_OBRA'], 'ValPref3',
             COL['SUMA_VALORES'], 'ValPref5'],
    'CAMPOS_USUARIO': ['GrpHRuta', 'CGH', 'Material', 'Ce.', COL['OP'],
                       'Indicador', 'clase de control', COL['NRO_PERSONAS'], COL['NRO_MAQUINAS']],
    'RECHAZO': ['GrPlf', COL['CLAVE_BUSQUEDA'], 'Material', 'Ce.', 'alternativa', 'alternativa',
                'posición', 'Relevancia', COL['PORCENTAJE_RECHAZO'],
                '% rechazo anterior', COL['DIFERENCIA'], 'Txt.brv.HRuta']
}

FINAL_COL_ORDER = [
    'GrpHRuta', 'CGH', 'Material', COL['CLAVE_BUSQUEDA'], 'Ce.', 'GrPlf', COL['OP'],
    COL['PORCENTAJE_RECHAZO'], COL['CANTIDAD_BASE'], COL['CANT_CALCULADA'],
    COL['DIFERENCIA'], COL['PESO_NETO'], COL['SECUENCIA'], 'ValPref', 'ValPref1',
    'ValPref2', COL['MANO_OBRA'], 'ValPref3', 'ValPref4', COL['SUMA_VALORES'],
    'ValPref5', 'Campo de usuario cantidad MANUAL', COL['NRO_PERSONAS'],
    'Campo de usuario cantidad MAQUINAS', COL['NRO_MAQUINAS'],
    'Texto breve operación', 'Ctrl', 'VerF', 'PstoTbjo', 'Cl.', 'Gr.fam.pre',
    'Texto breve de material', 'Txt.brv.HRuta', 'Bloq.vers.fabric.', 'Campo usuario unidad',
    'Campo usuario unidad.1', 'Cantidad', 'Contador', 'InBo', 'InBo.1', 'InBo.2',
    'Unnamed: 31', 'I'
]

COLUMNAS_A_SUMAR = ['ValPref', 'ValPref1', COL['MANO_OBRA'], 'ValPref3']

# --- 2. FUNCIONES DE LÓGICA ABSTRAÍDA ---

def _obtener_nombre_columna(cols: list, idx: int, default_name: str) -> str:
    return cols[idx] if idx < len(cols) else default_name

def _mapear_df(df_origen: pd.DataFrame, df_mapa: pd.DataFrame, col_clave_origen: str, col_clave_mapa: str, col_valor_mapa: str, col_destino: str, keep_mode: str = 'first'):
    mapa_series = (
        df_mapa.sort_values(by=col_valor_mapa, ascending=(keep_mode == 'first'), na_position='last')
        .drop_duplicates(subset=[col_clave_mapa], keep=keep_mode)
        .set_index(col_clave_mapa)[col_valor_mapa]
    )
    df_origen[col_destino] = df_origen[col_clave_origen].map(mapa_series)

def detectar_y_marcar_cantidad_atipica(grupo: pd.DataFrame) -> pd.Series:
    valores_no_nan = grupo[COL['CANT_CALCULADA']].dropna()
    if valores_no_nan.empty:
        return pd.Series(False, index=grupo.index)
    moda = Counter(valores_no_nan).most_common(1)[0][0]
    return grupo[COL['CANT_CALCULADA']] != moda

def filtrar_operaciones_impares_desde_31(df: pd.DataFrame) -> pd.DataFrame:
    if COL['OP'] not in df.columns:
        return pd.DataFrame()
    df_temp = df.copy()
    op_num = pd.to_numeric(df_temp[COL['OP']].astype(str).str.strip(), errors='coerce')
    condicion = (op_num.notna()) & (op_num >= 31) & (op_num % 2 != 0)
    return df_temp[condicion]

def obtener_secuencia(puesto_trabajo: str, df_secuencias: pd.DataFrame) -> Union[int, float]:
    psttbjo_str = str(puesto_trabajo).strip()
    try:
        for col_idx in range(df_secuencias.shape[1]):
            col_data = df_secuencias.iloc[:, col_idx].dropna()
            col_data_str = col_data.astype(str).str.strip()
            if psttbjo_str in set(col_data_str):
                return col_idx + 1
    except Exception:
        return np.nan
    return np.nan

# --- 3. FUNCIÓN DE CARGA Y LIMPIEZA SIMPLIFICADA ---

def cargar_y_limpiar_datos(file_original: io.BytesIO, file_info_externa: io.BytesIO, nombre_horno: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, str]]:
    config = HORNOS_CONFIG[nombre_horno]
    hoja_principal = config['HOJA_PRINCIPAL']
    cols_original = pd.read_excel(file_original, sheet_name=hoja_principal, nrows=0).columns.tolist()
    file_original.seek(0)
    cols_pn = pd.read_excel(file_original, sheet_name='Peso neto', nrows=0).columns.tolist()
    file_original.seek(0)
    cols_externo = pd.read_excel(file_info_externa, sheet_name='Especif y Rutas', nrows=0).columns.tolist()
    file_info_externa.seek(0)
    
    col_names = {
        'cant_base_leida': _obtener_nombre_columna(cols_original, IDX['CANTIDAD_BASE_LEIDA'], COL['CANTIDAD_BASE']),
        'material': _obtener_nombre_columna(cols_original, IDX['MATERIAL'], 'Material'),
        'psttbjo': _obtener_nombre_columna(cols_original, IDX['PSTTBJO'], 'PstoTbjo'),
        'grplf': _obtener_nombre_columna(cols_original, IDX['GRPLF'], 'GrPlf'),
        'material_pn': _obtener_nombre_columna(cols_pn, IDX['MATERIAL_PN'], 'Material'),
        'peso_neto_valor': _obtener_nombre_columna(cols_pn, IDX['PESO_NETO_VALOR'], 'Peso neto'),
        'nombre_col_rechazo_externa': _obtener_nombre_columna(cols_externo, IDX['RECHAZO_EXTERNA'], 'Columna AC'),
        'hoja_principal': hoja_principal
    }

    df_original = pd.read_excel(file_original, sheet_name=hoja_principal)
    file_original.seek(0)
    df_peso_neto = pd.read_excel(file_original, sheet_name='Peso neto')
    file_original.seek(0)
    df_secuencias = pd.read_excel(file_original, sheet_name=COL['HOJA_SALIDA_SECUENCIAS'])
    file_original.seek(0)
    columnas_mano_obra = [0, 1, 2, 3, 4]
    df_mano_obra = pd.read_excel(file_original, sheet_name=COL['HOJA_MANO_OBRA'], header=None,
                                 usecols=range(len(columnas_mano_obra)), names=columnas_mano_obra)
    file_original.seek(0)
    cols_a_leer_externo = [COL['CLAVE_EXTERNA'], COL['CANT_EXTERNA'], col_names['nombre_col_rechazo_externa']]
    df_externo = pd.read_excel(file_info_externa, sheet_name='Especif y Rutas', usecols=cols_a_leer_externo)
    file_info_externa.seek(0)
    
    rename_map = {col_names['cant_base_leida']: COL['CANTIDAD_BASE'], col_names['material']: 'Material', col_names['psttbjo']: 'PstoTbjo', col_names['grplf']: 'GrPlf'}
    df_original.rename(columns={k: v for k, v in rename_map.items() if k in df_original.columns}, inplace=True)
    return df_original, df_externo, df_peso_neto, df_secuencias, df_mano_obra, col_names

# --- 4. FUNCIÓN DE CREACIÓN DE HOJAS DE EXCEL (MODIFICADA PARA VINCULACIÓN TOTAL) ---

def crear_y_guardar_hoja(wb, df_base: pd.DataFrame, nombre_hoja: str, columnas_destino: list, fill_encabezado: PatternFill, font_negrita: Font, hoja_salida_name: str = None):
    """Crea una hoja donde cada celda es una fórmula que referencia a la hoja principal procesada."""
    df_temp = filtrar_operaciones_impares_desde_31(df_base) if nombre_hoja == COL['HOJA_SALIDA_CAMPOS_USUARIO'] else df_base.copy()
    
    if nombre_hoja in wb.sheetnames: del wb[nombre_hoja]
    ws = wb.create_sheet(nombre_hoja)
    ws.append(columnas_destino)
    
    if not df_temp.empty and hoja_salida_name:
        # Identificar las letras de columna en la hoja de origen
        mapeo_columnas_origen = {}
        for col in columnas_destino:
            if col in df_base.columns:
                idx = list(df_base.columns).index(col) + 1
                mapeo_columnas_origen[col] = get_column_letter(idx)

        # Escribir filas (fórmulas o valores fijos)
        for r_idx, (df_idx, _) in enumerate(df_temp.iterrows(), start=2):
            fila_origen = df_idx + 2 # +2 por encabezado y base 0 de pandas
            
            for c_idx, col_nombre in enumerate(columnas_destino, start=1):
                # Valores determinados/fijos solicitados
                if nombre_hoja == COL['HOJA_SALIDA_CAMPOS_USUARIO']:
                    if col_nombre == 'Indicador': 
                        ws.cell(row=r_idx, column=c_idx, value='x')
                        continue
                    elif col_nombre == 'clase de control': 
                        ws.cell(row=r_idx, column=c_idx, value='ZPP0006')
                        continue

                # Vinculación por fórmula dinámica
                if col_nombre in mapeo_columnas_origen:
                    letra_orig = mapeo_columnas_origen[col_nombre]
                    referencia = f"'{hoja_salida_name}'!{letra_orig}{fila_origen}"
                    ws.cell(row=r_idx, column=c_idx, value=f"=IF({referencia}=0,\"\",{referencia})")

    # Aplicar formato a encabezados
    indices_a_formatear = [columnas_destino.index(col) + 1 for col in COL['RESALTAR'] if col in columnas_destino]
    for col_idx in indices_a_formatear:
        ws.cell(row=1, column=col_idx).fill = fill_encabezado
        ws.cell(row=1, column=col_idx).font = font_negrita

# --- 5. FUNCIÓN PRINCIPAL DE PROCESAMIENTO REFACTORIZADA ---

def automatizacion_final_diferencia_reforzada(file_original: io.BytesIO, file_info_externa: io.BytesIO, nombre_horno: str) -> Tuple[bool, Union[str, io.BytesIO]]:
    config = HORNOS_CONFIG[nombre_horno]
    HOJA_SALIDA = config['HOJA_SALIDA']

    try:
        st.subheader(f"Preparando datos para **{nombre_horno}**... 📊")
        df_original, df_externo, df_peso_neto, df_secuencias, df_mano_obra, col_names = cargar_y_limpiar_datos(file_original, file_info_externa, nombre_horno)
        
        def limpiar(serie: pd.Series) -> pd.Series:
            return serie.astype(str).str.strip().str.replace(r'\W+', '', regex=True)

        df_original[COL['CLAVE_BUSQUEDA']] = limpiar(df_original['Material']) + limpiar(df_original['GrPlf']) + limpiar(df_original['PstoTbjo'])
        df_externo[COL['CLAVE_EXTERNA']] = limpiar(df_externo[COL['CLAVE_EXTERNA']])
        
        columna_para_secuencia = 'PstoTbjo'
        if COL['LINEA'] in df_original.columns and limpiar(df_original[COL['LINEA']]).str.len().gt(0).any():
            linea_limpia = df_original[COL['LINEA']].astype(str).str.strip()
            psttbjo_limpio = df_original['PstoTbjo'].astype(str).str.strip()
            df_original['PstoTbjo_Concat'] = np.where(linea_limpia.str.lower().isin(['nan', 'none', '', '-']), psttbjo_limpio, psttbjo_limpio + linea_limpia)
            columna_para_secuencia = 'PstoTbjo_Concat'
            
        df_original[COL['SECUENCIA']] = [obtener_secuencia(x, df_secuencias) for x in df_original[columna_para_secuencia]]

        mapa_max_cantidad = (df_externo.sort_values(by=COL['CANT_EXTERNA'], ascending=False, na_position='last')
                            .drop_duplicates(subset=[COL['CLAVE_EXTERNA']], keep='first').set_index(COL['CLAVE_EXTERNA'])[COL['CANT_EXTERNA']])
        df_original[COL['CANT_CALCULADA']] = df_original[COL['CLAVE_BUSQUEDA']].map(mapa_max_cantidad)
        _mapear_df(df_original, df_externo, COL['CLAVE_BUSQUEDA'], COL['CLAVE_EXTERNA'], col_names['nombre_col_rechazo_externa'], COL['PORCENTAJE_RECHAZO'])
        _mapear_df(df_original, df_peso_neto, 'Material', col_names['material_pn'], col_names['peso_neto_valor'], COL['PESO_NETO'])

        COL_PSTTBJO_MO, COL_TIEMPO_MO, COL_CANTIDAD_MAQUINAS_MO, COL_CANTIDAD_PERSONAS_MO = 0, 2, 3, 4
        df_mano_obra[COL_PSTTBJO_MO] = df_mano_obra[COL_PSTTBJO_MO].astype(str).str.strip()
        df_mano_obra['Calculo_MO_Tiempo'] = pd.to_numeric(df_mano_obra[COL_TIEMPO_MO], errors='coerce') * 60
        indices_terminan_en_1 = df_original[COL['OP']].astype(str).str.strip().str.endswith('1')
        psttbjo_filtrado = df_original.loc[indices_terminan_en_1, 'PstoTbjo'].astype(str).str.strip()
        
        map_mo_tiempo = df_mano_obra.drop_duplicates(subset=[COL_PSTTBJO_MO]).set_index(COL_PSTTBJO_MO)['Calculo_MO_Tiempo']
        map_personas = df_mano_obra.drop_duplicates(subset=[COL_PSTTBJO_MO]).set_index(COL_PSTTBJO_MO)[COL_CANTIDAD_PERSONAS_MO]
        map_maquinas = df_mano_obra.drop_duplicates(subset=[COL_PSTTBJO_MO]).set_index(COL_PSTTBJO_MO)[COL_CANTIDAD_MAQUINAS_MO]
        
        df_original.loc[indices_terminan_en_1, COL['MANO_OBRA']] = psttbjo_filtrado.map(map_mo_tiempo)
        df_original.loc[indices_terminan_en_1, COL['NRO_PERSONAS']] = psttbjo_filtrado.map(map_personas)
        df_original.loc[indices_terminan_en_1, COL['NRO_MAQUINAS']] = psttbjo_filtrado.map(map_maquinas)
        df_original[COL['SUMA_VALORES']] = np.nan

        cant_base_float = pd.to_numeric(df_original[COL['CANTIDAD_BASE']].astype(str).str.replace(',', '.', regex=False).str.strip(), errors='coerce')
        df_original[COL['CANTIDAD_BASE']] = np.trunc(cant_base_float)
        df_original[COL['DIFERENCIA']] = np.nan
        
        atipicos_series = df_original.groupby([COL['PESO_NETO'], COL['SECUENCIA']], dropna=True).apply(detectar_y_marcar_cantidad_atipica)
        idx_original = atipicos_series.index.get_level_values(-1)
        df_original[COL['ATIPICO']] = pd.Series(atipicos_series.values, index=idx_original).reindex(df_original.index, fill_value=False)

        df_original = df_original.drop(columns=['PstoTbjo_Concat']) if 'PstoTbjo_Concat' in df_original.columns else df_original
        df_original_final = df_original.reindex(columns=[c for c in FINAL_COL_ORDER if c in df_original.columns])

        file_original.seek(0)
        wb = load_workbook(file_original)
        fill_anomalia = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        fill_encabezado = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        font_negrita = Font(bold=True)
        
        if HOJA_SALIDA in wb.sheetnames: del wb[HOJA_SALIDA]
        ws = wb.create_sheet(HOJA_SALIDA)
        for row in dataframe_to_rows(df_original_final, header=True, index=False): ws.append(row)

        col_dif_idx = df_original_final.columns.get_loc(COL['DIFERENCIA']) + 1
        col_base_letter = get_column_letter(df_original_final.columns.get_loc(COL['CANTIDAD_BASE']) + 1)
        col_calculada_letter = get_column_letter(df_original_final.columns.get_loc(COL['CANT_CALCULADA']) + 1)
        col_suma_valores_idx = df_original_final.columns.get_loc(COL['SUMA_VALORES']) + 1
        col_sum_letters = [get_column_letter(df_original_final.columns.get_loc(col) + 1) for col in COLUMNAS_A_SUMAR if col in df_original_final.columns]

        for r in range(2, len(df_original_final) + 2):
            ws.cell(row=r, column=col_dif_idx, value=f'=ROUNDDOWN({col_base_letter}{r}, 0) - {col_calculada_letter}{r}').number_format = '#,##0.00'
            sum_expr = f'SUM({",".join([f"{l}{r}" for l in col_sum_letters])})'
            ws.cell(row=r, column=col_suma_valores_idx, value=f'=IF({sum_expr}=0,"",{sum_expr})').number_format = '#,##0.00'
            if df_original.iloc[r-2][COL['ATIPICO']]:
                ws.cell(row=r, column=df_original_final.columns.get_loc(COL['CANT_CALCULADA']) + 1).fill = fill_anomalia
                
        # Creación de Hojas Vinculadas dinámicamente
        crear_y_guardar_hoja(wb, df_original_final, COL['HOJA_SALIDA_LSMW'], COLUMNAS_OUTPUT['LSMW'], fill_encabezado, font_negrita, HOJA_SALIDA)
        crear_y_guardar_hoja(wb, df_original_final, COL['HOJA_SALIDA_CAMPOS_USUARIO'], COLUMNAS_OUTPUT['CAMPOS_USUARIO'], fill_encabezado, font_negrita, HOJA_SALIDA)
        crear_y_guardar_hoja(wb, df_original_final, COL['HOJA_SALIDA_RECHAZO'], COLUMNAS_OUTPUT['RECHAZO'], fill_encabezado, font_negrita, HOJA_SALIDA)
        
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        return True, output_buffer

    except Exception as e:
        return False, f"❌ Error: {e}"

# --- 6. INTERFAZ DE STREAMLIT ---

def main():
    st.set_page_config(page_title="Automatización Hornos", layout="centered")
    st.title("⚙️ Automatización Verificación de datos - HORNOS")
    hornos_disponibles = list(HORNOS_CONFIG.keys())
    selected_horno = st.radio("**1. Seleccione el Horno:**", hornos_disponibles, index=4, horizontal=True)
    
    col1, col2 = st.columns(2)
    with col1:
        file_original = st.file_uploader("Carga la base de datos original", type=['xlsx'])
    with col2:
        file_externa = st.file_uploader("Carga el archivo externo", type=['xlsb', 'xlsx'])

    if st.button(f"▶️ PROCESAR {selected_horno}", type="primary", use_container_width=True):
        if file_original and file_externa:
            success, resultado = automatizacion_final_diferencia_reforzada(io.BytesIO(file_original.getvalue()), io.BytesIO(file_externa.getvalue()), selected_horno)
            if success:
                st.success(f"✅ Completado.")
                st.warning("⚠️ Abre el archivo y presiona **F9** para activar las fórmulas.")
                st.download_button("⬇️ Descargar", data=resultado, file_name=f"Reporte_{selected_horno}.xlsx", use_container_width=True)
            else:
                st.error(resultado)

if __name__ == "__main__":
    main()





